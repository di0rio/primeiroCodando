from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

def criar_planilha():
    """Cria uma nova planilha do Excel com as colunas necessárias."""
    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = 'Nome da Empresa'
    sheet['B1'] = 'Cidade'
    sheet['C1'] = 'Estado'
    sheet['D1'] = 'Telefone'
    sheet['E1'] = 'Email'
    sheet['F1'] = 'Nome do Contato'
    sheet['G1'] = 'Comprou?'
    sheet['H1'] = 'Data Agendamento'
    wb.save('clientes.xlsx')

def cadastrar_cliente(nome_empresa, cidade, estado, telefone, email, contato, comprou):
    """Cadastra um novo cliente na planilha."""
    try:
        wb = load_workbook('clientes.xlsx')
        sheet = wb.active
        sheet.append([nome_empresa, cidade, estado, telefone, email, contato, comprou, None])
        wb.save('clientes.xlsx')
        print("Cliente cadastrado com sucesso!")
    except FileNotFoundError:
        print("A planilha 'clientes.xlsx' não foi encontrada. Crie a planilha primeiro.")

def agendar_ligacao(nome_empresa, data):
    """Agenda uma ligação para um cliente existente."""
    try:
        wb = load_workbook('clientes.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == nome_empresa:
                row[7].value = data
                wb.save('clientes.xlsx')
                print(f"Ligação agendada para {nome_empresa} em {data.strftime('%d/%m/%Y')}.")
                return
        print(f"Cliente {nome_empresa} não encontrado.")
    except FileNotFoundError:
        print("A planilha 'clientes.xlsx' não foi encontrada. Crie a planilha primeiro.")

def gerar_relatorio():
    """Exibe um relatório dos clientes com ligações agendadas."""
    try:
        wb = load_workbook('clientes.xlsx')
        sheet = wb.active
        print("Clientes com ligações agendadas:")
        for row in sheet.iter_rows(min_row=2):
            if row[7].value is not None:
                print(f"Nome da Empresa: {row[0].value}, Data Agendamento: {row[7].value.strftime('%d/%m/%Y')}")
    except FileNotFoundError:
        print("A planilha 'clientes.xlsx' não foi encontrada. Crie a planilha primeiro.")

if __name__ == "__main__":
    # Cria a planilha caso não exista
    try:
        load_workbook('clientes.xlsx')
    except FileNotFoundError:
        criar_planilha()

    while True:
        print("\n--- Menu ---")
        print("1. Cadastrar Cliente")
        print("2. Agendar Ligação")
        print("3. Gerar Relatório")
        print("4. Sair")

        opcao = input("Digite a opção desejada: ")

        if opcao == '1':
            nome_empresa = input("Nome da Empresa: ")
            cidade = input("Cidade: ")
            estado = input("Estado: ")
            telefone = input("Telefone: ")
            email = input("Email: ")
            contato = input("Nome do Contato: ")
            comprou = input("Cliente comprou? (Sim/Não): ").lower() == 'sim'
            cadastrar_cliente(nome_empresa, cidade, estado, telefone, email, contato, comprou)
        elif opcao == '2':
            nome_empresa = input("Nome da Empresa: ")
            data_str = input("Data da Ligação (dd/mm/aaaa): ")
            try:
                data = datetime.datetime.strptime(data_str, '%d/%m/%Y').date()
                agendar_ligacao(nome_empresa, data)
            except ValueError:
                print("Formato de data inválido. Utilize dd/mm/aaaa.")
        elif opcao == '3':
            gerar_relatorio()
        elif opcao == '4':
            break
        else:
            print("Opção inválida.")